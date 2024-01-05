import os
import json
import tensorflow as tf
from PIL import Image
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns
import io

# Model directory and paths
MODEL_DIR = "D:/code_test/Darkspots TensorFlow"  # Update with the actual path to your model
TEST_DIR = "D:/code_test/test"
VALIDATION_DIR = "D:/code_test/validation"
OUTPUT_DIR = "D:/code_test"
NEW_PIC = "D:/code_test/Real_people"

# Load the model using tf.saved_model.load
class Model(object):
    def __init__(self, model_dir):
        self.model = tf.saved_model.load(model_dir)
        self.signature = self.model.signatures["serving_default"]

    def predict(self, image: Image.Image):
        image = np.asarray(image)
        image = (image / 255.0).astype(np.float32)
        image = tf.image.resize(image, (224, 224))  # Adjust the size as needed

        # Wrap the image in a batch dimension
        image = tf.expand_dims(image, 0)

        # Make the prediction using the actual output tensor names
        result = self.signature(tf.constant(image))  # Pass the batched image
        predictions = result['Prediction'].numpy()  # Use the correct output tensor name for predictions
        confidences = result['Confidences'].numpy()  # Use the correct output tensor name for confidences

        # Extract the class index and confidence
        class_index = np.argmax(confidences)
        confidence = confidences[0][class_index]

        return class_index, confidence

# Function to create a confusion matrix plot and add it to the Excel sheet
def add_confusion_matrix_plot(ws, predictions, image_types, sheet_name):
    # Create a confusion matrix
    confusion_matrix = np.zeros((len(image_types), len(image_types)), dtype=int)
    for i, true_label in enumerate(image_types):
        true_label_indices = np.where(predictions['true_labels'] == i)[0]
        for index in true_label_indices:
            predicted_label = predictions['predicted_labels'][index]
            confusion_matrix[i][predicted_label] += 1

    # Create a heatmap plot of the confusion matrix
    plt.figure(figsize=(8, 6))
    sns.set(font_scale=1.2)
    sns.heatmap(confusion_matrix, annot=True, fmt="d", cmap="Blues",
                xticklabels=image_types, yticklabels=image_types)
    plt.xlabel("Predicted Labels")
    plt.ylabel("True Labels")
    plt.title(f"Confusion Matrix ({sheet_name})")

    # Save the plot as an image in memory
    plot_image = io.BytesIO()
    plt.savefig(plot_image, format="png", bbox_inches="tight")
    plot_image.seek(0)

    # Add the plot image to the Excel sheet
    img = openpyxl.drawing.image.Image(plot_image)
    img.width = 400
    img.height = 300
    ws.add_image(img, 'E2')  # Adjust the cell location as needed

# Function to predict the image and store results in Excel
def predict_and_create_excel(model, image_dir, excel_path, sheet_name, calculate_confusion_matrix=True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws['A1'] = "File Name"
    ws['B1'] = "Image Type"
    ws['C1'] = "Prediction"
    ws['D1'] = "Confidence"

    image_types = os.listdir(image_dir)
    row = 2

    # Create arrays to store true and predicted labels for the confusion matrix
    true_labels = []
    predicted_labels = []

    for image_type_index, image_type in enumerate(image_types):
        type_path = os.path.join(image_dir, image_type)
        if os.path.isdir(type_path):
            for image_file in os.listdir(type_path):
                if image_file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                    image_path = os.path.join(type_path, image_file)
                    image = Image.open(image_path)
                    if image.mode != "RGB":
                        image = image.convert("RGB")

                    class_index, confidence = model.predict(image)
                    prediction = "Darkspots" if class_index == 0 else "Healthy_skin"
                    ws[f'A{row}'] = image_file
                    ws[f'B{row}'] = image_type
                    ws[f'C{row}'] = prediction
                    ws[f'D{row}'] = confidence
                    row += 1

                    # Append true and predicted labels for confusion matrix
                    true_labels.append(image_type_index)
                    predicted_labels.append(class_index)

    # Create a dictionary to store true and predicted labels
    predictions = {'true_labels': np.array(true_labels), 'predicted_labels': np.array(predicted_labels)}

    if calculate_confusion_matrix:
        # Add the confusion matrix plot to the Excel sheet
        add_confusion_matrix_plot(ws, predictions, image_types, sheet_name)

    wb.save(excel_path)

if __name__ == "__main__":
    model = Model(MODEL_DIR)

    # Predict and create Excel for the test set with confusion matrix
    test_excel_path = os.path.join(OUTPUT_DIR, "test_predictions.xlsx")
    predict_and_create_excel(model, TEST_DIR, test_excel_path, "Test")

    # Predict and create Excel for the validation set with confusion matrix
    validation_excel_path = os.path.join(OUTPUT_DIR, "validation_predictions.xlsx")
    predict_and_create_excel(model, VALIDATION_DIR, validation_excel_path, "Validation")
    
    # Predict and create Excel for the new images without confusion matrix
    new_images_excel_path = os.path.join(OUTPUT_DIR, "new_images_predictions.xlsx")
    predict_and_create_excel(model, NEW_PIC, new_images_excel_path, "New Images", calculate_confusion_matrix=False)